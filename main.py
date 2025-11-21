import os
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Dict, Any
from datetime import datetime, timedelta
import csv
from io import StringIO, BytesIO
from openpyxl import load_workbook

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def read_root():
    return {"message": "Hello from FastAPI Backend!"}

@app.get("/api/hello")
def hello():
    return {"message": "Hello from the backend API!"}

@app.get("/test")
def test_database():
    """Test endpoint to check if database is available and accessible"""
    response = {
        "backend": "✅ Running",
        "database": "❌ Not Available",
        "database_url": None,
        "database_name": None,
        "connection_status": "Not Connected",
        "collections": []
    }
    
    try:
        from database import db
        if db is not None:
            response["database"] = "✅ Available"
            response["database_url"] = "✅ Configured"
            response["database_name"] = db.name if hasattr(db, 'name') else "✅ Connected"
            response["connection_status"] = "Connected"
            try:
                collections = db.list_collection_names()
                response["collections"] = collections[:10]
                response["database"] = "✅ Connected & Working"
            except Exception as e:
                response["database"] = f"⚠️  Connected but Error: {str(e)[:50]}"
        else:
            response["database"] = "⚠️  Available but not initialized"
    except ImportError:
        response["database"] = "❌ Database module not found (run enable-database first)"
    except Exception as e:
        response["database"] = f"❌ Error: {str(e)[:50]}"
    
    response["database_url"] = "✅ Set" if os.getenv("DATABASE_URL") else "❌ Not Set"
    response["database_name"] = "✅ Set" if os.getenv("DATABASE_NAME") else "❌ Not Set"
    return response

# ---------------------- Optimization Endpoint ----------------------

def _norm(s: str) -> str:
    return (s or "").strip().lower()


def read_rows_from_csv(data: bytes) -> List[Dict[str, Any]]:
    text = data.decode('utf-8', errors='ignore')
    reader = csv.DictReader(StringIO(text))
    return [dict(row) for row in reader]


def read_rows_from_excel(data: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        d: Dict[str, Any] = {}
        for i, val in enumerate(r):
            key = headers[i] if i < len(headers) else f"col{i}"
            d[key] = val if val is not None else ""
        out.append(d)
    return out


def parse_demand_rows(rows: List[Dict[str, Any]], city: str) -> Dict[datetime, int]:
    demand_map: Dict[datetime, int] = {}
    city = (city or "").strip()
    for row in rows:
        try:
            keys = { _norm(k): k for k in row.keys() }
            # City
            city_val = (str(row.get(keys.get('city_code')) or row.get(keys.get('city')) or "")).strip()
            if city and city_val != city:
                continue
            # Timestamp
            raw_ts = (
                row.get(keys.get('slot_started_local_at'))
                or row.get(keys.get('timestamp'))
                or row.get(keys.get('time'))
                or row.get(keys.get('datetime'))
            )
            if raw_ts is None or raw_ts == "":
                continue
            # Excel may give datetime objects already
            if isinstance(raw_ts, datetime):
                dt = raw_ts
            else:
                s = str(raw_ts).replace('Z', '+00:00')
                if 'T' not in s and ' ' in s:
                    s = s.replace(' ', 'T')
                dt = datetime.fromisoformat(s)
            # Demand
            demand_raw = (
                row.get(keys.get('final_order_forecast'))
                or row.get(keys.get('available_capacity'))
                or row.get(keys.get('demand'))
                or 0
            )
            if isinstance(demand_raw, (int, float)):
                demand = int(demand_raw)
            else:
                demand = int(float(str(demand_raw).strip() or '0'))
            demand_map[dt] = demand
        except Exception:
            continue
    return demand_map


def parse_riders_rows(rows: List[Dict[str, Any]], city: str) -> List[Dict[str, Any]]:
    riders: List[Dict[str, Any]] = []
    city = (city or "").strip()
    for row in rows:
        try:
            keys = { _norm(k): k for k in row.keys() }
            row_city = (str(row.get(keys.get('ciudad')) or row.get(keys.get('city')) or "")).strip()
            if city and row_city != city:
                continue
            rider_id = str(row.get(keys.get('rider id')) or row.get(keys.get('rider_id')) or row.get(keys.get('id')) or "").strip()
            start_raw = row.get(keys.get('available from')) or row.get(keys.get('start')) or row.get(keys.get('start_time'))
            end_raw = row.get(keys.get('available to')) or row.get(keys.get('end')) or row.get(keys.get('end_time'))
            if not start_raw or not end_raw:
                continue
            if isinstance(start_raw, datetime):
                start_dt = start_raw
            else:
                s = str(start_raw).replace('Z', '+00:00')
                if 'T' not in s and ' ' in s:
                    s = s.replace(' ', 'T')
                start_dt = datetime.fromisoformat(s)
            if isinstance(end_raw, datetime):
                end_dt = end_raw
            else:
                e = str(end_raw).replace('Z', '+00:00')
                if 'T' not in e and ' ' in e:
                    e = e.replace(' ', 'T')
                end_dt = datetime.fromisoformat(e)
            if end_dt <= start_dt:
                continue
            riders.append({'rider_id': rider_id, 'start': start_dt, 'end': end_dt})
        except Exception:
            continue
    return riders


def daterange(start: datetime, end: datetime, step_minutes: int = 30) -> List[datetime]:
    out = []
    cur = start
    while cur <= end:
        out.append(cur)
        cur = cur + timedelta(minutes=step_minutes)
    return out


def overlap(a_start: datetime, a_end: datetime, b_start: datetime, b_end: datetime) -> bool:
    return a_start < b_end and b_start < a_end


@app.post("/api/optimize")
async def optimize(
    demand_file: UploadFile = File(...),
    riders_file: UploadFile = File(...),
    start_date: str = Form(...),
    end_date: str = Form(...),
    city: str = Form(...),
    interval_minutes: int = Form(30)
):
    """
    Accepts two files (CSV or Excel) and returns coverage series.
    Demand supported headers:
      - city_code, slot_started_local_at, final_order_forecast (preferred)
      - or city, timestamp/time, demand
    Riders supported headers:
      - CIUDAD, RIDER ID, Available from, Available to (preferred)
      - or city, rider_id, start, end
    Dates accept 'YYYY-MM-DD HH:MM' or ISO 8601.
    """
    try:
        start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00').replace(' ', 'T'))
        end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00').replace(' ', 'T'))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date format. Use 'YYYY-MM-DD HH:MM'.")

    if end_dt <= start_dt:
        raise HTTPException(status_code=400, detail="end_date must be after start_date")

    # Read files into rows
    try:
        d_bytes = await demand_file.read()
        r_bytes = await riders_file.read()
    except Exception:
        raise HTTPException(status_code=400, detail="Unable to read uploaded files")

    def read_rows(upload: UploadFile, data: bytes) -> List[Dict[str, Any]]:
        name = (upload.filename or '').lower()
        try:
            if name.endswith('.csv'):
                return read_rows_from_csv(data)
            # assume Excel
            return read_rows_from_excel(data)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse file {upload.filename}: {e}")

    d_rows = read_rows(demand_file, d_bytes)
    r_rows = read_rows(riders_file, r_bytes)

    demand_map = parse_demand_rows(d_rows, city.strip())
    riders = parse_riders_rows(r_rows, city.strip())

    # Build timeline
    times = daterange(start_dt, end_dt, interval_minutes)

    series: List[Dict[str, Any]] = []
    total_unmet = 0
    total_surplus = 0

    for i in range(len(times)):
        t0 = times[i]
        t1 = times[i] + timedelta(minutes=interval_minutes)
        demand = demand_map.get(t0, 0)
        available = sum(1 for r in riders if overlap(r['start'], r['end'], t0, t1))
        staffed = min(demand, available)
        unmet = max(demand - available, 0)
        surplus = max(available - demand, 0)
        total_unmet += unmet
        total_surplus += surplus
        series.append({
            'time': t0.isoformat(),
            'demand': demand,
            'available': available,
            'staffed': staffed,
            'unmet': unmet,
            'surplus': surplus
        })

    summary = {
        'interval_minutes': interval_minutes,
        'points': len(series),
        'total_unmet': total_unmet,
        'total_surplus': total_surplus,
        'city': city,
        'start': start_dt.isoformat(),
        'end': end_dt.isoformat(),
    }

    return {"summary": summary, "series": series}


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
